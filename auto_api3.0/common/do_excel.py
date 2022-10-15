# -*-coding:utf-8-*-
# @Time     :2022-09-24 20:42
# @Author   :Digua
# 此模块用于测试数据的读取
import json
import os

import openpyxl

from test_cases.case import Case
from common import contants as cts
from common.config import ReadConf


class DoExcel:
    def __init__(self, file_name, sheet_name, setting_name='wms'):
        """

        :param file_name: 测试用例文件名
        :param sheet_name: 测试用例所在表名
        :param setting_name: 该测试文件所使用的配置文件
        """
        self.file_name = os.path.join(cts.case_data_path, file_name)
        self.sheet_name = sheet_name
        self.setting_name = setting_name
        # 实例化配置文件读取器
        # 默认读取“test_config.ini”配置文件
        read_conf = ReadConf()
        self.conf_info = read_conf.get_section_conf(self.setting_name)
        # 加载目标xls文件
        self.wb = openpyxl.load_workbook(self.file_name)
        # 获取表的内容
        self.sheet = self.wb[self.sheet_name]

    def read_case(self):

        # 定义存放结果Case对象的临时列表
        case_temp = []
        for i in range(2, self.sheet.max_row + 1):
            # 实例化一个Case对象，用于集成用例数据
            case = Case()
            # 获取用例的行,列
            case.case_row_num = i
            case.case_test_result_col = self.conf_info.get("case_test_result")
            case.case_act_result_col = self.conf_info.get("case_act_result")
            # 获取case_id
            case.case_id = self.__get_value(i, "case_id")
            case.case_module = self.__get_value(i,"case_module")
            case.case_title = self.__get_value(i, "case_title")
            case.case_method = self.__get_value(i, "case_method")
            case.case_url =self.__get_value(i, "case_url")
            case.case_headers = self.__get_value(i, "case_headers")
            case.case_data_template = self.__get_value(i, "case_data_template")
            # 按照对应模板更新data
            # if case.case_data_template:
            #     with open(os.path.join(cts.case_data_path,"data_template.json"),
            #               'r',encoding='utf-8') as f:
            #         data_template = json.load(f)[case.case_data_template]
            #         data_template.update(case.case_data)
            #         case.case_data=data_template
            case.case_data =self.__get_value(i, "case_data")
            if case.case_data:
            # 对data数据 进行处理
                case.case_data = eval('{' + case.case_data.replace('\n', ',') + '}')
            case.case_expected = self.__get_value(i, "case_expected")
            case.case_storage_name = self.__get_value(i,"case_storage_name")
            case_temp.append(case)

        return case_temp

    def __get_value(self, number, case_filed):
        # 用于处理配置文件对应字段为None的情况
        number2 = self.conf_info.get(case_filed,None)
        if number2 == None:
            return number2
        else:
            result = self.sheet.cell(number,
                                   int(self.conf_info.get(case_filed,None))).value
            return result

    def write(self,row,col,value):
        """

        :param row: 指定的行
        :param col: 指定的列
        :param value: 指定的值
        :return:
        """
        wb = openpyxl.load_workbook(self.file_name)
        sheet = wb[self.sheet_name]
        sheet.cell(row,col).value = value
        wb.save(self.file_name)
        # self.wb.close()

    def write_case_result(self, case):
        self.write(case.case_row_num,
                   int(case.case_test_result_col),
                   case.case_test_result)


if __name__ == '__main__':
    read_excel = DoExcel("wms_case.xlsx","库区管理","wms")
    data = read_excel.read_case()
    data2 = data[0].case_data
    print(data2)
    # data3 = {'name': '接口测试收货', 'typeName': '收货暂存区', 'typeCode': 1, 'area': '1', 'nihao':00}
    # data2.update(data3)
    # print(data2)
    # print(data2)
    # wordlist = '{' + data2.replace('\n', ',') + '}'
    # print(wordlist)
    # data3 =  eval(wordlist)
    # print(type(data3.get("typeCode")))
    # print(str_turn_dict_mobile(data2))
    # read_excel.write(2, 1, "喜喜")
    # data2 = read_excel.read_case()
    # # print(data)
    # for i in data:
    #     print(i.case_row_num)
    #     print(i.case_data)
    #
    # # read_excel.write(2, 1, "你好")
    # data2 = read_excel.read_case()
    # # print(data)
    # for i in data2:
    #     print(i.case_id)
    #
    # read_excel.write(2, 1, "旺仔")
    # data3 = read_excel.read_case()
    # # print(data)
    # for i in data3:
    #     print(i.case_id)

    # read_excel.close()

