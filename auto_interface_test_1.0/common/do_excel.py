# -*-coding:utf-8-*-
# @FileName :do_excel.py
# @Time     :2022-07-21 16:13
# @Author   :Digua
"""
用于读写excel文件中的测试用例
"""
import os

import openpyxl

from common import contants as cf
from common.config import Read_Conf


class Case:
    """
    用于集成获取到的测试用例数据
    """

    def __init__(self):
        self.case_id = None
        self.case_title = None
        self.english_Name = None
        self.case_url = None
        self.case_header = None
        self.case_data = None
        self.case_method = None
        self.case_expected = None


class DoExcel:
    """
    用于从excel文件获取测试数据
    """

    def __init__(self, file_name, sheet_name='库区管理', section='wms'):
        """

        :param file_name: 需要操作的文件名称
        :param sheet_name: 需要操作的表名称
        :param section: 对应的配置文件选项
        """
        self.file_name = os.path.join(cf.datas_path, file_name)
        self.sheet_name = sheet_name
        self.section = section

    @property  # 私有方法，仅可通过 [实例.read] 方式访问，无法修改
    def read(self) -> [Case, ]:
        # 实例化读取配置文件得对象
        read_excel = Read_Conf("test.ini")
        # 获取目标文件的book对象
        wb = openpyxl.load_workbook(self.file_name)
        # 根据表名称获取工作表对象
        sheet = wb[self.sheet_name]
        # 定义存放需要反对的Case对象列表
        tmp = []
        # 从第二行开始遍历，以便跳过表头数据
        for i in range(2, sheet.max_row + 1):
            # 实例化一个Case，以便收集测试数据
            case = Case()
            # 获取测试用例id
            case.case_id = sheet.cell(i,
                                      int(read_excel.get_option
                                          (self.section, "case_id")
                                          )
                                      ).value
            # 获取测试用例标题
            case.case_title = sheet.cell(i,
                                         int(read_excel.get_option
                                             (self.section, "case_title")
                                             )
                                         ).value
            # 获取测试用例url
            case.case_url = read_excel.get_option(self.section, 'base_url')\
                            + sheet.cell(i,
                                         int(read_excel.get_option
                                             (self.section, "case_url")
                                             )).value  # 拼接url
            # 获取测试用例数据
            case.case_data = sheet.cell(i,
                                        int(read_excel.get_option
                                            (self.section, "case_data"))
                                        ).value
            # 获取测试用例请求方式
            case.case_method = sheet.cell(i,
                                          int(read_excel.get_option
                                              (self.section, "case_method"))
                                          ).value
            # 获取测试用例期望值
            case.case_expected = sheet.cell(i,
                                            int(read_excel.get_option
                                                (self.section, "case_expected")
                                                )
                                            ).value
            tmp.append(case)
        return tmp

    def write(self, row, col, value):
        """向excel指定行、列中写数据"""
        # 加载目标文件
        wb = openpyxl.load_workbook(self.file_name)
        # 根据指定的表名获取工作表
        sheet = wb[self.sheet_name]
        # 修改指定的单元格数据
        sheet.cell(row, col).value = value
        # 保存修改后的数据
        wb.save(self.file_name)
        # 关闭工作空间
        wb.close()


if __name__ == '__main__':
    from datas.get_case import wms_case
    cases = wms_case["库区管理"]
    print(cases)
    # data = DoExcel("wms_case.xlsx")
    # data2 = data.read
    for case_1 in cases:
        print(case_1.case_id)
        print(case_1.case_title)
        print(case_1.case_url)
        print(case_1.case_data)
        print(case_1.case_method)
        print(case_1.case_expected)
